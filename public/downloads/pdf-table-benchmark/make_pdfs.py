"""
検証用PDFの生成。同じ表を2つの経路で作る。

  A: 表計算ソフト（openpyxl で .xlsx → LibreOffice で PDF 変換）
     … 日本の実務資料の作られ方として最も多い経路
  B: Webブラウザ（HTML → Chromium の印刷機能で PDF）
     … 業務システムからの帳票出力を模した経路

同じ中身でも作り方が違えばPDFの内部構造は変わる。
抽出精度が「表の中身」で決まるのか「作り方」で決まるのかを切り分けるため、
2経路それぞれで測る。
"""
import os, subprocess, shutil, html, asyncio, sys
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from tables import TABLES

OUT = "/root/lab/pdf/pdfs"
WORK = "/root/lab/pdf/work"
LANDSCAPE = {"T08"}


def disp_width(s):
    """全角を2、半角を1として見た目の幅を数える"""
    return sum(2 if ord(c) > 0x2E80 else 1 for c in str(s).replace("\n", ""))


def col_widths(t):
    ws = []
    for c in range(len(t["rows"][0])):
        w = max(disp_width(r[c]) if c < len(r) else 0 for r in t["rows"])
        ws.append(min(max(w + 4, 8), 34))
    return ws


# ---------------------------------------------------------------- A: xlsx 経路
def build_xlsx(t, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    thin = Side(style="thin", color="000000")
    if t["border"] == "grid":
        bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    elif t["border"] == "horizontal":
        bd = Border(top=thin, bottom=thin)
    else:
        bd = Border()

    for ri, row in enumerate(t["rows"], start=1):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(ri, ci)
            cell.value = val            # すべて文字列として入れる（自動書式で表示が変わるのを防ぐ）
            cell.number_format = "@"
            cell.border = bd
            cell.alignment = Alignment(
                wrap_text=True, vertical="center",
                horizontal="left" if ci == 1 else "right",
            )
            if ri == 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                if t.get("header_fill"):
                    cell.fill = PatternFill("solid", fgColor="DDE7E2")

    for (r0, c0, c1) in t.get("merges", []):
        ws.merge_cells(start_row=r0 + 1, start_column=c0 + 1, end_row=r0 + 1, end_column=c1 + 1)

    for ci, w in enumerate(col_widths(t), start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    # 折り返しが起きる行は高さを確保する（切れると正解と比較できなくなるため）
    for ri, row in enumerate(t["rows"], start=1):
        lines = max(str(v).count("\n") + 1 for v in row)
        longest = max(disp_width(v) for v in row)
        if lines > 1 or longest > 30:
            ws.row_dimensions[ri].height = 18 * max(lines, 2 if longest > 30 else 1) + 8

    ws.page_setup.orientation = "landscape" if t["id"] in LANDSCAPE else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    wb.save(path)


def xlsx_to_pdf(xlsx, outdir):
    subprocess.run(
        ["soffice", "--headless", "--norestore", "--nologo",
         "-env:UserInstallation=file:///tmp/loprof_pdf",
         "--convert-to", "pdf:calc_pdf_Export", "--outdir", outdir, xlsx],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=180,
    )


# ---------------------------------------------------------------- B: HTML 経路
def build_html(t, path):
    if t["border"] == "grid":
        css_cell = "border:1px solid #000;"
    elif t["border"] == "horizontal":
        css_cell = "border-top:1px solid #000; border-bottom:1px solid #000;"
    else:
        css_cell = "border:0;"

    widths = col_widths(t)
    total = sum(widths)
    colgroup = "".join(f'<col style="width:{w/total*100:.2f}%">' for w in widths)

    merges = {(r0, c0): c1 - c0 + 1 for (r0, c0, c1) in t.get("merges", [])}
    skip = set()
    for (r0, c0, c1) in t.get("merges", []):
        for c in range(c0 + 1, c1 + 1):
            skip.add((r0, c))

    trs = []
    for ri, row in enumerate(t["rows"]):
        tds = []
        for ci, v in enumerate(row):
            if (ri, ci) in skip:
                continue
            span = merges.get((ri, ci))
            tag = "th" if ri == 0 else "td"
            attr = f' colspan="{span}"' if span else ""
            align = "left" if ci == 0 else "right"
            if ri == 0:
                align = "center"
            body = html.escape(str(v)).replace("\n", "<br>")
            tds.append(f'<{tag}{attr} style="text-align:{align}">{body}</{tag}>')
        trs.append("<tr>" + "".join(tds) + "</tr>")

    size = "A4 landscape" if t["id"] in LANDSCAPE else "A4 portrait"
    fill = "background:#dde7e2;" if t.get("header_fill") else ""
    doc = f"""<!doctype html><html><head><meta charset="utf-8"><style>
@page {{ size:{size}; margin:16mm; }}
body {{ font-family:'Noto Sans CJK JP',sans-serif; font-size:10.5pt; margin:0; }}
table {{ border-collapse:collapse; width:100%; table-layout:fixed; }}
th,td {{ {css_cell} padding:4px 6px; vertical-align:middle; word-break:break-all; }}
th {{ font-weight:700; {fill} }}
</style></head><body><table>{colgroup}{''.join(trs)}</table></body></html>"""
    open(path, "w", encoding="utf-8").write(doc)


async def html_to_pdf(pairs):
    from playwright.async_api import async_playwright
    async with async_playwright() as p:
        b = await p.chromium.launch(
            executable_path="/opt/pw-browsers/chromium-1194/chrome-linux/chrome",
            args=["--no-sandbox"])
        page = await b.new_page()
        for src, dst, land in pairs:
            await page.goto("file://" + src, wait_until="networkidle")
            await page.pdf(path=dst, format="A4", landscape=land,
                           margin={"top": "16mm", "bottom": "16mm",
                                   "left": "16mm", "right": "16mm"},
                           print_background=True)
        await b.close()


def main():
    shutil.rmtree(OUT, ignore_errors=True)
    shutil.rmtree(WORK, ignore_errors=True)
    os.makedirs(OUT, exist_ok=True)
    os.makedirs(WORK, exist_ok=True)

    for t in TABLES:
        x = f"{WORK}/{t['id']}.xlsx"
        build_xlsx(t, x)
        xlsx_to_pdf(x, WORK)
        src = f"{WORK}/{t['id']}.pdf"
        if os.path.exists(src):
            shutil.move(src, f"{OUT}/{t['id']}_calc.pdf")
        else:
            print(f"  !! {t['id']} calc変換に失敗")

    pairs = []
    for t in TABLES:
        h = f"{WORK}/{t['id']}.html"
        build_html(t, h)
        pairs.append((h, f"{OUT}/{t['id']}_html.pdf", t["id"] in LANDSCAPE))
    asyncio.run(html_to_pdf(pairs))

    made = sorted(os.listdir(OUT))
    print(f"生成したPDF: {len(made)}本")
    for f in made:
        print("  ", f, f"{os.path.getsize(OUT+'/'+f)//1024}KB")


if __name__ == "__main__":
    main()
