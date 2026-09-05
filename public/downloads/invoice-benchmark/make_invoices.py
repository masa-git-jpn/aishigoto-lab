#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
検証用の請求書PDFを50枚生成し、正解データ(truth.json)を書き出す。

5つのテンプレート x 10枚。実務で実際に届く請求書のばらつきを再現する。
  A: HTML -> Chromium印刷 / 明細表に格子罫線
  B: HTML -> Chromium印刷 / 横罫線のみ（縦線なし）
  C: HTML -> Chromium印刷 / 罫線なし（レイアウトだけで表を作る）
  D: xlsx -> LibreOffice変換 / 格子罫線
  E: xlsx -> LibreOffice変換 / 横罫線のみ・値引き行(△表記)・明細20行でページまたぎ

依存: reportlab不要 / openpyxl, chromium, libreoffice
"""
import json, os, random, subprocess, sys, shutil
from pathlib import Path
from datetime import date, timedelta

ROOT = Path(__file__).parent
PDF = ROOT / "pdfs"
WORK = ROOT / "work"
CHROMIUM = "/opt/pw-browsers/chromium"

random.seed(20260905)

VENDORS = [
    ("株式会社ミドリ電機", "東京都千代田区神田錦町1-2-3"),
    ("有限会社さくら工業", "大阪府大阪市北区梅田4-5-6"),
    ("合同会社ブルースカイ", "神奈川県横浜市西区北幸2-1-8"),
    ("株式会社アオイ商事", "愛知県名古屋市中区栄3-9-1"),
    ("トウホク精機株式会社", "宮城県仙台市青葉区一番町2-4-7"),
    ("株式会社ひかりソリューションズ", "福岡県福岡市博多区博多駅前1-1-1"),
    ("西川製作所株式会社", "京都府京都市下京区烏丸通6-3"),
    ("株式会社コスモテック", "北海道札幌市中央区大通西5-8"),
    ("ヤマト梱包資材株式会社", "埼玉県さいたま市大宮区桜木町1-7-5"),
    ("株式会社ニシキ印刷", "広島県広島市中区八丁堀12-2"),
]

ITEMS = [
    ("保守サポート費（月額）", "式"), ("サーバー利用料", "式"),
    ("設計作業費", "人日"), ("部品A-2200", "個"),
    ("梱包資材（Lサイズ）", "箱"), ("出張旅費", "式"),
    ("ライセンス更新料", "本"), ("印刷代（A4カラー）", "枚"),
    ("運送費", "回"), ("追加開発費", "人日"),
    ("定期点検作業", "回"), ("消耗品一式", "式"),
]

def yen(n):
    return f"{n:,}"

def zenkaku(s):
    tbl = str.maketrans("0123456789,", "０１２３４５６７８９，")
    return s.translate(tbl)


def build_record(i, tmpl):
    """1枚分のデータと正解を作る"""
    vendor, addr = VENDORS[i % len(VENDORS)]
    d = date(2026, 7, 1) + timedelta(days=random.randint(0, 60))
    nrow = 20 if tmpl == "E" else random.randint(3, 7)
    rows = []
    for _ in range(nrow):
        name, unit = random.choice(ITEMS)
        qty = random.randint(1, 30)
        price = random.choice([800, 1200, 2500, 4800, 12000, 35000, 48000, 120000])
        rows.append({"name": name, "qty": qty, "unit": unit,
                     "price": price, "amount": qty * price})
    discount = None
    if tmpl == "E" and random.random() < 0.5:
        discount = -random.choice([5000, 10000, 20000])
        rows.append({"name": "お値引き", "qty": 1, "unit": "式",
                     "price": discount, "amount": discount})
    subtotal = sum(r["amount"] for r in rows)
    tax = int(subtotal * 0.1)
    total = subtotal + tax
    # 請求書番号の書式もばらけさせる（実務ではベンダごとに違う）
    nofmt = i % 4
    if nofmt == 0:
        no = f"INV-2026-{i+1:04d}"
    elif nofmt == 1:
        no = f"{d.strftime('%Y%m%d')}-{i+1:03d}"
    elif nofmt == 2:
        no = f"No.A{i+1:05d}"
    else:
        no = f"SEI-{d.strftime('%y%m')}-{i+1:03d}"
    # 日付の書式もばらけさせる
    dfmt = i % 3
    if dfmt == 0:
        dstr = f"{d.year}年{d.month}月{d.day}日"
    elif dfmt == 1:
        dstr = d.strftime("%Y/%m/%d")
    else:
        dstr = d.strftime("%Y-%m-%d")
    return {
        "file": f"invoice_{i+1:02d}_{tmpl}.pdf",
        "template": tmpl,
        "invoice_no": no, "invoice_no_raw": no,
        "issue_date": d.isoformat(), "issue_date_raw": dstr,
        "vendor": vendor, "vendor_addr": addr,
        "subtotal": subtotal, "tax": tax, "total": total,
        "rows": rows,
        "n_rows": len(rows),
    }


CSS_COMMON = """
@page { size: A4; margin: 18mm 16mm; }
body { font-family: 'Noto Sans CJK JP', sans-serif; font-size: 10.5pt; color:#000; }
h1 { font-size: 20pt; letter-spacing: 8pt; text-align:center; margin: 0 0 14pt; }
.meta { text-align: right; font-size: 10pt; line-height: 1.7; }
.to { font-size: 13pt; border-bottom: 1px solid #000; display:inline-block; padding: 0 30pt 3pt 0; margin: 10pt 0 14pt; }
.from { text-align: right; font-size: 9.5pt; line-height:1.6; margin-top: 6pt;}
.total-box { margin: 12pt 0 16pt; font-size: 13pt; }
table { width: 100%; border-collapse: collapse; font-size: 10pt; }
th, td { padding: 4pt 6pt; }
th { background: #f0f0f0; }
td.num, th.num { text-align: right; }
.sum { width: 46%; margin-left: auto; margin-top: 10pt; }
"""
CSS_A = "table, th, td { border: 0.6pt solid #333; }"
CSS_B = "th, td { border-bottom: 0.6pt solid #333; } table {border-top:0.6pt solid #333;}"
CSS_C = "th { background: none; border: none; } td { border: none; }"


def html_for(rec, css_variant):
    r = rec
    body_rows = "\n".join(
        f"<tr><td>{x['name']}</td><td class='num'>{x['qty']}</td><td>{x['unit']}</td>"
        f"<td class='num'>{yen(x['price'])}</td><td class='num'>{yen(x['amount'])}</td></tr>"
        for x in r["rows"])
    return f"""<!doctype html><html lang="ja"><head><meta charset="utf-8">
<style>{CSS_COMMON}{css_variant}</style></head><body>
<h1>請求書</h1>
<div class="meta">請求書番号：{r['invoice_no']}<br>発行日：{r['issue_date_raw']}</div>
<div class="to">株式会社サンプル商会 御中</div>
<div class="from">{r['vendor']}<br>{r['vendor_addr']}<br>登録番号 T{random.randint(1000000000000,9999999999999)}</div>
<div class="total-box">下記のとおりご請求申し上げます。<br>
<b>ご請求金額（税込）　{yen(r['total'])} 円</b></div>
<table>
<thead><tr><th>品目</th><th class="num">数量</th><th>単位</th><th class="num">単価</th><th class="num">金額</th></tr></thead>
<tbody>{body_rows}</tbody></table>
<table class="sum">
<tr><td>小計</td><td class="num">{yen(r['subtotal'])}</td></tr>
<tr><td>消費税(10%)</td><td class="num">{yen(r['tax'])}</td></tr>
<tr><td><b>合計</b></td><td class="num"><b>{yen(r['total'])}</b></td></tr>
</table>
<p style="margin-top:18pt;font-size:9.5pt">お振込先：サンプル銀行 本店営業部 普通 1234567<br>
お支払期限：翌月末日</p>
</body></html>"""


def make_html_pdf(rec, css_variant):
    WORK.mkdir(exist_ok=True)
    h = WORK / (rec["file"].replace(".pdf", ".html"))
    h.write_text(html_for(rec, css_variant), encoding="utf-8")
    out = PDF / rec["file"]
    subprocess.run([CHROMIUM, "--headless", "--disable-gpu", "--no-sandbox",
                    "--no-pdf-header-footer", f"--print-to-pdf={out}", h.as_uri()],
                   check=True, capture_output=True, timeout=120)


def make_xlsx_pdf(rec, grid):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    WORK.mkdir(exist_ok=True)
    wb = Workbook(); ws = wb.active; ws.title = "請求書"
    thin = Side(style="thin", color="000000")
    if grid:
        bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    else:
        bd = Border(bottom=thin)
    ws.column_dimensions['A'].width = 34
    for c in "BCDE":
        ws.column_dimensions[c].width = 12
    ws["A1"] = "請  求  書"; ws["A1"].font = Font(size=18, bold=True)
    ws["D2"] = "請求書番号"; ws["E2"] = rec["invoice_no"]
    ws["D3"] = "発行日"; ws["E3"] = rec["issue_date_raw"]
    ws["A3"] = "株式会社サンプル商会 御中"; ws["A3"].font = Font(size=12)
    ws["A5"] = rec["vendor"]; ws["A6"] = rec["vendor_addr"]
    ws["A8"] = "ご請求金額（税込）"
    tot = yen(rec["total"]) + " 円"
    if rec["template"] == "E":
        tot = zenkaku(yen(rec["total"])) + " 円"   # 全角数字の請求書は実在する
    ws["C8"] = tot; ws["C8"].font = Font(size=13, bold=True)
    hdr = 10
    for j, name in enumerate(["品目", "数量", "単位", "単価", "金額"]):
        c = ws.cell(row=hdr, column=j + 1, value=name)
        c.font = Font(bold=True); c.border = bd
        c.alignment = Alignment(horizontal="center")
    r = hdr + 1
    for x in rec["rows"]:
        amt = x["amount"]; price = x["price"]
        # 値引きは会計慣習の △ 表記にする
        pv = f"△{yen(abs(price))}" if price < 0 else yen(price)
        av = f"△{yen(abs(amt))}" if amt < 0 else yen(amt)
        vals = [x["name"], x["qty"], x["unit"], pv, av]
        for j, v in enumerate(vals):
            c = ws.cell(row=r, column=j + 1, value=v)
            c.border = bd
            if j in (1, 3, 4):
                c.alignment = Alignment(horizontal="right")
        r += 1
    r += 1
    for label, v in [("小計", rec["subtotal"]), ("消費税(10%)", rec["tax"]), ("合計", rec["total"])]:
        ws.cell(row=r, column=4, value=label)
        c = ws.cell(row=r, column=5, value=yen(v))
        c.alignment = Alignment(horizontal="right")
        if label == "合計":
            c.font = Font(bold=True)
        r += 1
    xp = WORK / rec["file"].replace(".pdf", ".xlsx")
    wb.save(xp)
    subprocess.run(["libreoffice", "--headless", "--convert-to", "pdf",
                    "--outdir", str(WORK), str(xp)], check=True,
                   capture_output=True, timeout=180)
    shutil.move(str(WORK / xp.name.replace(".xlsx", ".pdf")), str(PDF / rec["file"]))


def main():
    PDF.mkdir(exist_ok=True)
    recs = []
    plan = ["A"] * 10 + ["B"] * 10 + ["C"] * 10 + ["D"] * 10 + ["E"] * 10
    for i, t in enumerate(plan):
        rec = build_record(i, t)
        recs.append(rec)
    for rec in recs:
        t = rec["template"]
        if t == "A":
            make_html_pdf(rec, CSS_A)
        elif t == "B":
            make_html_pdf(rec, CSS_B)
        elif t == "C":
            make_html_pdf(rec, CSS_C)
        elif t == "D":
            make_xlsx_pdf(rec, grid=True)
        else:
            make_xlsx_pdf(rec, grid=False)
        print("made", rec["file"], flush=True)
    (ROOT / "truth.json").write_text(
        json.dumps(recs, ensure_ascii=False, indent=1), encoding="utf-8")
    print("total", len(recs))


if __name__ == "__main__":
    main()
