#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
請求書PDFのフォルダを丸ごと読んで、1つのExcelにまとめる。

  python3 extract_invoices.py ./pdfs -o invoices.xlsx

出力Excel
  シート「一覧」   : 1行 = 1請求書（番号・発行日・請求元・小計・消費税・合計・要確認）
  シート「明細」   : 1行 = 明細1行（どのファイルの何行目かが分かる）
  シート「要確認」 : 検算が合わなかったファイルだけを抜き出したもの

依存: pdfplumber（表抽出の比較用）と poppler-utils の pdftotext
      pip install pdfplumber openpyxl
      apt install poppler-utils
"""
import argparse, json, re, subprocess, sys, unicodedata
from pathlib import Path

# ---------------------------------------------------------------- 文字の正規化
# pdfminer系（pdfplumber）は一部の漢字を「部首」のコードポイントで返すことがある。
# 例: 西(U+897F) → ⻄(U+2EC4 CJK RADICAL WEST TWO)
# NFKC はこのブロック(U+2E80-U+2EF3)を変換してくれないので、自前で表を作る。
def _radical_map():
    kangxi = {}
    for cp in range(0x2F00, 0x2FD6):          # 康熙部首（NFKCで漢字になる）
        ch = chr(cp)
        try:
            n = unicodedata.name(ch)
        except ValueError:
            continue
        kangxi[n.replace("KANGXI RADICAL ", "")] = unicodedata.normalize("NFKC", ch)
    suffix = re.compile(r" (ONE|TWO|THREE|FOUR|FIVE|SIX)$")
    m = {}
    for cp in range(0x2E80, 0x2EF4):          # CJK部首補助（NFKCが効かない）
        ch = chr(cp)
        try:
            n = unicodedata.name(ch)
        except ValueError:
            continue
        base = n.replace("CJK RADICAL ", "")
        for cand in (base, suffix.sub("", base),
                     re.sub(r"^(C-SIMPLIFIED|J-SIMPLIFIED|SIMPLIFIED) ", "", suffix.sub("", base))):
            if cand in kangxi:
                m[ch] = kangxi[cand]
                break
    return m

RADICALS = _radical_map()

def normalize(s: str) -> str:
    """全角→半角、部首→漢字、マイナス表記の統一。抽出の一番最初に必ず通す。"""
    if not s:
        return ""
    s = "".join(RADICALS.get(c, c) for c in s)
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("△", "-").replace("▲", "-").replace("−", "-").replace("―", "-")
    return s

NUM = re.compile(r"-?[\d,]+")

def to_int(s):
    """'1,234' '-5,000' '１２３' を int に。数字が無ければ None。"""
    if s is None:
        return None
    m = NUM.search(normalize(str(s)).replace(" ", ""))
    if not m:
        return None
    try:
        return int(m.group(0).replace(",", ""))
    except ValueError:
        return None

# ---------------------------------------------------------------- テキスト取得
def text_poppler(path: Path) -> str:
    r = subprocess.run(["pdftotext", "-layout", str(path), "-"],
                       capture_output=True, text=True, encoding="utf-8")
    return normalize(r.stdout)

def text_pdfplumber(path: Path) -> str:
    import pdfplumber
    with pdfplumber.open(path) as pdf:
        return normalize("\n".join((p.extract_text(layout=True) or "") for p in pdf.pages))

# ---------------------------------------------------------------- ヘッダ項目
DATE_PATTERNS = [
    (re.compile(r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日"), (1, 2, 3)),
    (re.compile(r"(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})"), (1, 2, 3)),
]

def find_date(text):
    for line in text.split("\n"):
        if "発行日" in line or "請求日" in line:
            for pat, idx in DATE_PATTERNS:
                m = pat.search(line)
                if m:
                    y, mo, d = (int(m.group(i)) for i in idx)
                    return f"{y:04d}-{mo:02d}-{d:02d}"
    for pat, idx in DATE_PATTERNS:      # 見つからなければ本文の最初の日付
        m = pat.search(text)
        if m:
            y, mo, d = (int(m.group(i)) for i in idx)
            return f"{y:04d}-{mo:02d}-{d:02d}"
    return None

NO_KEY = re.compile(r"(請求書?番号|請求番号|INVOICE\s*NO\.?|伝票番号)[:：\s]*")

def find_invoice_no(text):
    for line in text.split("\n"):
        m = NO_KEY.search(line)
        if m:
            rest = line[m.end():].strip()
            if rest:
                return rest.split()[0]
    return None

CORP = re.compile(r"(株式会社|有限会社|合同会社|合資会社)")

def find_vendor(text):
    """請求元＝会社名を含む行のうち、宛先（御中/様）ではないもの。"""
    cands = []
    for line in text.split("\n"):
        s = line.strip()
        if not s or "御中" in s or s.endswith("様"):
            continue
        if CORP.search(s):
            cands.append(s.split()[-1] if len(s.split()) == 1 else s.strip())
    return cands[0] if cands else None

def find_amount(text, keys, exclude=()):
    for line in text.split("\n"):
        if any(k in line for k in keys) and not any(x in line for x in exclude):
            nums = NUM.findall(line.replace(" ", ""))
            nums = [n for n in nums if n not in ("10", "8")]      # 「(10%)」を拾わない
            if nums:
                return to_int(nums[-1])
    return None

# ---------------------------------------------------------------- 明細行
HDR = re.compile(r"品\s*目|品\s*名|摘\s*要|内\s*容")
END = re.compile(r"小\s*計|合\s*計|消費税")
ROW = re.compile(
    r"^(?P<name>.*?\S)\s+"
    r"(?P<qty>-?[\d,]+)\s*"
    r"(?P<unit>[^\s\d,.-]{0,4}?)\s+"
    r"(?P<price>-?[\d,]+)\s+"
    r"(?P<amount>-?[\d,]+)\s*$")

def find_rows(text):
    rows, inside = [], False
    for raw in text.split("\n"):
        line = raw.rstrip()
        if not line.strip():
            continue
        if not inside:
            if HDR.search(line) and ("金額" in line or "数量" in line):
                inside = True
            continue
        if END.search(line):
            inside = False
            continue
        m = ROW.match(line.strip())
        if not m:
            continue
        amt = to_int(m.group("amount"))
        if amt is None:
            continue
        rows.append({
            "name": m.group("name").strip(),
            "qty": to_int(m.group("qty")),
            "unit": m.group("unit").strip(),
            "price": to_int(m.group("price")),
            "amount": amt,
        })
    return rows

# ---------------------------------------------------------------- 1ファイル処理
def extract(path: Path, engine="poppler"):
    text = text_poppler(path) if engine == "poppler" else text_pdfplumber(path)
    rows = find_rows(text)
    rec = {
        "file": path.name,
        "invoice_no": find_invoice_no(text),
        "issue_date": find_date(text),
        "vendor": find_vendor(text),
        "subtotal": find_amount(text, ["小計"]),
        "tax": find_amount(text, ["消費税", "税額"]),
        "total": find_amount(text, ["合計"], exclude=["小計"])
                 or find_amount(text, ["ご請求金額", "請求金額"]),
        "rows": rows,
    }
    rec["checks"] = reconcile(rec)
    return rec

def reconcile(rec):
    """検算。ここで引っかかったファイルだけ人間が見ればよい、という状態にする。"""
    ng = []
    for k in ("invoice_no", "issue_date", "vendor", "total"):
        if not rec.get(k):
            ng.append(f"{k}が取れていない")
    if not rec["rows"]:
        ng.append("明細が0行")
    s = sum(r["amount"] for r in rec["rows"]) if rec["rows"] else None
    if s is not None and rec.get("subtotal") is not None and s != rec["subtotal"]:
        ng.append(f"明細合計({s:,})と小計({rec['subtotal']:,})が不一致")
    if (rec.get("subtotal") is not None and rec.get("tax") is not None
            and rec.get("total") is not None
            and rec["subtotal"] + rec["tax"] != rec["total"]):
        ng.append("小計+消費税が合計と一致しない")
    for r in rec["rows"]:
        if r["qty"] is not None and r["price"] is not None and r["amount"] is not None:
            if r["qty"] * r["price"] != r["amount"]:
                ng.append(f"数量x単価が金額と合わない行がある（{r['name']}）")
                break
    return ng

# ---------------------------------------------------------------- Excel出力
def write_excel(recs, out):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active; ws.title = "一覧"
    head = ["ファイル", "請求書番号", "発行日", "請求元", "小計", "消費税", "合計", "明細行数", "要確認"]
    ws.append(head)
    for r in recs:
        ws.append([r["file"], r["invoice_no"], r["issue_date"], r["vendor"],
                   r["subtotal"], r["tax"], r["total"], len(r["rows"]),
                   " / ".join(r["checks"])])
    ws2 = wb.create_sheet("明細")
    ws2.append(["ファイル", "行", "品目", "数量", "単位", "単価", "金額"])
    for r in recs:
        for i, x in enumerate(r["rows"], 1):
            ws2.append([r["file"], i, x["name"], x["qty"], x["unit"], x["price"], x["amount"]])
    ws3 = wb.create_sheet("要確認")
    ws3.append(["ファイル", "内容"])
    for r in recs:
        if r["checks"]:
            ws3.append([r["file"], " / ".join(r["checks"])])
    warn = PatternFill("solid", fgColor="FFF2CC")
    for sheet in (ws, ws2, ws3):
        for c in sheet[1]:
            c.font = Font(bold=True)
        sheet.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        if row[8].value:
            for c in row:
                c.fill = warn
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["I"].width = 46
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["C"].width = 26
    wb.save(out)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("folder")
    ap.add_argument("-o", "--out", default="invoices.xlsx")
    ap.add_argument("--engine", choices=["poppler", "pdfplumber"], default="poppler")
    ap.add_argument("--json", default=None)
    a = ap.parse_args()
    files = sorted(Path(a.folder).glob("*.pdf"))
    recs = []
    for p in files:
        try:
            recs.append(extract(p, a.engine))
        except Exception as e:                       # 1枚落ちても全体は止めない
            recs.append({"file": p.name, "invoice_no": None, "issue_date": None,
                         "vendor": None, "subtotal": None, "tax": None, "total": None,
                         "rows": [], "checks": [f"読み取りに失敗: {e}"]})
    write_excel(recs, a.out)
    if a.json:
        Path(a.json).write_text(json.dumps(recs, ensure_ascii=False, indent=1),
                                encoding="utf-8")
    ng = sum(1 for r in recs if r["checks"])
    print(f"{len(recs)}枚を読み込み、{ng}枚が要確認。→ {a.out}")

if __name__ == "__main__":
    main()
