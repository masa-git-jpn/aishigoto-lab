#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
請求書PDFのフォルダを丸ごと読んで、1つのExcelにまとめる（v2）。

  python3 extract_invoices_v2.py ./pdfs -o invoices.xlsx

v1（extract_invoices.py）を想定外の請求書50枚に当てて壊れたところを直したもの。
直した点は4つ。
  1. テキスト層の有無を最初に判定する（スキャンPDFを「0円」として静かに通さない）
  2. 明細は「表として」取りにいき、取れなければテキスト行に落とす
     → 格子罫線があるPDFでは、セル内で折り返した品目名が復元できる
  3. 明細の列数と列名の揺れを許容する（数量・単価が無い／税率の列がある／英語）
  4. 小計・消費税が複数行あるとき（8%と10%の混在）は合算する

出力Excel
  シート「一覧」   : 1行 = 1請求書
  シート「明細」   : 1行 = 明細1行
  シート「要確認」 : 検算に引っかかったファイルだけ

依存: pip install pdfplumber openpyxl / apt install poppler-utils
"""
import argparse, json, re, subprocess, unicodedata
from pathlib import Path

# ---------------------------------------------------------------- 文字の正規化
# pdfminer系（pdfplumber）は一部の漢字を「部首」のコードポイントで返すことがある。
# 例: 西(U+897F) → ⻄(U+2EC4)。NFKC はこのブロックを変換しないので自前で表を作る。
def _radical_map():
    kangxi = {}
    for cp in range(0x2F00, 0x2FD6):
        ch = chr(cp)
        try:
            n = unicodedata.name(ch)
        except ValueError:
            continue
        kangxi[n.replace("KANGXI RADICAL ", "")] = unicodedata.normalize("NFKC", ch)
    suffix = re.compile(r" (ONE|TWO|THREE|FOUR|FIVE|SIX)$")
    m = {}
    for cp in range(0x2E80, 0x2EF4):
        ch = chr(cp)
        try:
            n = unicodedata.name(ch)
        except ValueError:
            continue
        base = n.replace("CJK RADICAL ", "")
        for cand in (base, suffix.sub("", base),
                     re.sub(r"^(C-SIMPLIFIED|J-SIMPLIFIED|SIMPLIFIED) ", "",
                            suffix.sub("", base))):
            if cand in kangxi:
                m[ch] = kangxi[cand]
                break
    return m

RADICALS = _radical_map()

# 表抽出の経路では、全角記号が「見た目が同じ別のコードポイント」で返ることがある。
# 例: ／(U+FF0F) → ∕(U+2215 DIVISION SLASH)。NFKC はこれも変換しない。
CONFUSABLES = {"\u2215": "/", "\u2044": "/", "\u2236": ":", "\u02d0": ":"}

def normalize(s):
    if not s:
        return ""
    s = "".join(CONFUSABLES.get(c, RADICALS.get(c, c)) for c in s)
    s = unicodedata.normalize("NFKC", s)
    return s.replace("△", "-").replace("▲", "-").replace("−", "-").replace("―", "-")

NUM = re.compile(r"-?[\d,]+")
PCT = re.compile(r"^-?\d+(\.\d+)?%$")

def to_int(s):
    if s is None:
        return None
    t = normalize(str(s)).replace(" ", "").replace("¥", "").replace("円", "")
    m = NUM.search(t)
    if not m:
        return None
    try:
        return int(m.group(0).replace(",", ""))
    except ValueError:
        return None

# ---------------------------------------------------------------- テキスト取得
def page_text(path):
    r = subprocess.run(["pdftotext", "-layout", str(path), "-"],
                       capture_output=True, text=True, encoding="utf-8")
    return normalize(r.stdout)

def has_text_layer(text, min_chars=40):
    """スキャンPDFはここで弾く。返り値 False なら OCR が要る。"""
    return len(re.sub(r"\s", "", text)) >= min_chars

# ---------------------------------------------------------------- ヘッダ項目
ERA = {"令和": 2018, "平成": 1988, "昭和": 1925}
DATE_PATTERNS = [
    (re.compile(r"(令和|平成|昭和)\s*(\d{1,2}|元)\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日"), "wareki"),
    (re.compile(r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日"), "ymd"),
    (re.compile(r"(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})"), "ymd"),
]
DATE_KEYS = ("発行日", "請求日", "作成日", "Date", "DATE", "Issue")

def _parse_date(line):
    for pat, kind in DATE_PATTERNS:
        m = pat.search(line)
        if not m:
            continue
        if kind == "wareki":
            era, yy, mo, d = m.groups()
            y = ERA[era] + (1 if yy == "元" else int(yy))
        else:
            y, mo, d = (int(x) for x in m.groups())
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
    return None

def find_date(text):
    for line in text.split("\n"):
        if any(k in line for k in DATE_KEYS):
            got = _parse_date(line)
            if got:
                return got
    return _parse_date(text)

NO_KEY = re.compile(r"(請求書?番号|請求番号|伝票番号|Invoice\s*No\.?|INVOICE\s*NO\.?|No\.)[:：\s]*",
                    re.IGNORECASE)

def find_invoice_no(text):
    for line in text.split("\n"):
        m = NO_KEY.search(line)
        if m:
            rest = line[m.end():].strip()
            if rest:
                return rest.split()[0]
    return None

CORP = re.compile(r"(株式会社|有限会社|合同会社|合資会社|\(株\)|Co\.,?\s*Ltd)")

def find_vendor(text):
    for line in text.split("\n"):
        s = line.strip()
        if not s or "御中" in s or s.endswith("様"):
            continue
        if CORP.search(s):
            return s
    return None

def _line_amount(line, drop=("%",)):
    nums = [n for n in NUM.findall(line.replace(" ", ""))
            if n not in ("10", "8", "0")]
    return to_int(nums[-1]) if nums else None

def find_amount(text, keys, exclude=(), mode="last"):
    """mode='sum' は、8%と10%で行が分かれている小計・消費税を足し合わせる。"""
    vals = []
    for line in text.split("\n"):
        if any(k in line for k in keys) and not any(x in line for x in exclude):
            v = _line_amount(line)
            if v is not None:
                vals.append(v)
    if not vals:
        return None
    return sum(vals) if mode == "sum" else vals[-1]

# ---------------------------------------------------------------- 明細（表として）
COL_KEYS = {
    "name": ("品目", "品名", "摘要", "内容", "項目", "description", "item"),
    "qty": ("数量", "個数", "qty", "quantity"),
    "unit": ("単位", "unit"),
    "price": ("単価", "unitprice", "price"),
    "amount": ("金額", "amount", "total"),
    "rate": ("税率", "rate", "tax"),
}
END_WORDS = ("小計", "合計", "消費税", "subtotal", "total", "tax")

def _col_kind(head):
    h = normalize(head or "").replace(" ", "").lower()
    if not h:
        return None
    # 「単価(Unit Price)」を「単位(Unit)」と取り違えないよう、price を先に見る
    for kind in ("qty", "price", "rate", "unit", "amount", "name"):
        if any(k in h for k in COL_KEYS[kind]):
            return kind
    return None

def rows_from_tables(path):
    """格子罫線があるPDFはここで取れる。セル内改行もつながった状態で返る。"""
    import pdfplumber
    out = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                hi = None
                for i, row in enumerate(table[:4]):
                    kinds = [_col_kind(c) for c in row]
                    if "name" in kinds and "amount" in kinds:
                        hi, cols = i, kinds
                        break
                if hi is None:
                    continue
                for row in table[hi + 1:]:
                    cells = [normalize((c or "").replace("\n", "")).strip() for c in row]
                    first = "".join(cells[:2])
                    if any(w in first.lower() for w in END_WORDS):
                        break
                    rec = {"name": "", "qty": None, "unit": "", "price": None, "amount": None}
                    for kind, val in zip(cols, cells):
                        if kind == "name" and val:
                            rec["name"] = (rec["name"] + val).strip()
                        elif kind == "unit":
                            rec["unit"] = val
                        elif kind in ("qty", "price", "amount"):
                            rec[kind] = to_int(val)
                    if rec["name"] and rec["amount"] is not None:
                        out.append(rec)
    return out

# ---------------------------------------------------------------- 明細（テキスト行）
HDR = re.compile(r"品\s*目|品\s*名|摘\s*要|内\s*容|description|item", re.IGNORECASE)
AMT = re.compile(r"金\s*額|amount", re.IGNORECASE)
END = re.compile(r"小\s*計|合\s*計|消費税|subtotal|total|tax", re.IGNORECASE)
SPLIT = re.compile(r"\s{2,}")

def rows_from_text(text):
    """罫線が無いPDF向け。列の区切りは「空白2つ以上」で見る。"""
    rows, inside = [], False
    for raw in text.split("\n"):
        line = raw.rstrip()
        if not line.strip():
            continue
        if not inside:
            if HDR.search(line) and AMT.search(line):
                inside = True
            continue
        if END.search(line):
            inside = False
            continue
        cells = [c.strip() for c in SPLIT.split(line.strip()) if c.strip()]
        if len(cells) < 2:
            continue
        name, rest = cells[0], cells[1:]
        if rest and PCT.match(rest[-1].replace(" ", "")):   # 税率の列は落とす
            rest = rest[:-1]
        vals = [to_int(c) for c in rest]
        if not vals or vals[-1] is None:
            continue
        rec = {"name": name, "qty": None, "unit": "", "price": None, "amount": vals[-1]}
        if len(vals) >= 2:
            rec["price"] = vals[-2]
        if len(vals) >= 3:
            rec["qty"] = vals[-3]
            m = re.search(r"\d+\s*([^\s\d,.-]+)", rest[-3])
            if m:
                rec["unit"] = m.group(1)
        if rec["name"]:
            rows.append(rec)
    return rows

# ---------------------------------------------------------------- 1ファイル処理
def extract(path):
    text = page_text(path)
    rec = {"file": path.name, "route": None, "invoice_no": None, "issue_date": None,
           "vendor": None, "subtotal": None, "tax": None, "total": None, "rows": []}
    if not has_text_layer(text):
        rec["route"] = "none"
        rec["checks"] = ["テキスト層が無い（スキャンPDF）。OCRが必要"]
        return rec
    rec.update({
        "invoice_no": find_invoice_no(text),
        "issue_date": find_date(text),
        "vendor": find_vendor(text),
        "subtotal": find_amount(text, ["小計", "Subtotal"], mode="sum"),
        "tax": find_amount(text, ["消費税", "税額", "Tax"], mode="sum"),
        "total": find_amount(text, ["合計", "Total"], exclude=["小計", "Subtotal"])
                 or find_amount(text, ["ご請求金額", "請求金額"]),
    })
    rows = rows_from_tables(path)          # まず表として取る
    rec["route"] = "table"
    if not rows:                            # 取れなければテキスト行に落とす
        rows = rows_from_text(text)
        rec["route"] = "text"
    rec["rows"] = rows
    rec["checks"] = reconcile(rec)
    return rec

def reconcile(rec):
    ng = []
    for k, label in (("invoice_no", "請求書番号"), ("issue_date", "発行日"),
                     ("vendor", "請求元"), ("total", "合計")):
        if not rec.get(k):
            ng.append(f"{label}が取れていない")
    if not rec["rows"]:
        ng.append("明細が0行")
    s = sum(r["amount"] for r in rec["rows"]) if rec["rows"] else None
    if s is not None and rec.get("subtotal") is not None and s != rec["subtotal"]:
        ng.append(f"明細合計({s:,})と小計({rec['subtotal']:,})が不一致")
    if (rec.get("subtotal") is not None and rec.get("tax") is not None
            and rec.get("total") is not None
            and rec["subtotal"] + rec["tax"] != rec["total"]):
        ng.append("小計+消費税が合計と一致しない")
    for r in rec["rows"]:
        if None not in (r["qty"], r["price"], r["amount"]) and r["qty"] * r["price"] != r["amount"]:
            ng.append(f"数量x単価が金額と合わない行がある（{r['name'][:14]}）")
            break
    return ng

# ---------------------------------------------------------------- Excel出力
def write_excel(recs, out):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook(); ws = wb.active; ws.title = "一覧"
    ws.append(["ファイル", "請求書番号", "発行日", "請求元", "小計", "消費税", "合計",
               "明細行数", "取得経路", "要確認"])
    for r in recs:
        ws.append([r["file"], r["invoice_no"], r["issue_date"], r["vendor"],
                   r["subtotal"], r["tax"], r["total"], len(r["rows"]),
                   r.get("route"), " / ".join(r["checks"])])
    ws2 = wb.create_sheet("明細")
    ws2.append(["ファイル", "行", "品目", "数量", "単位", "単価", "金額"])
    for r in recs:
        for i, x in enumerate(r["rows"], 1):
            ws2.append([r["file"], i, x["name"], x["qty"], x["unit"], x["price"], x["amount"]])
    ws3 = wb.create_sheet("要確認")
    ws3.append(["ファイル", "内容"])
    for r in recs:
        if r["checks"]:
            ws3.append([r["file"], " / ".join(r["checks"])])
    warn = PatternFill("solid", fgColor="FFF2CC")
    for sheet in (ws, ws2, ws3):
        for c in sheet[1]:
            c.font = Font(bold=True)
        sheet.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        if row[9].value:
            for c in row:
                c.fill = warn
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["J"].width = 46
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["C"].width = 40
    wb.save(out)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("folder")
    ap.add_argument("-o", "--out", default="invoices.xlsx")
    ap.add_argument("--json", default=None)
    a = ap.parse_args()
    recs = []
    for p in sorted(Path(a.folder).glob("*.pdf")):
        try:
            recs.append(extract(p))
        except Exception as e:
            recs.append({"file": p.name, "route": "error", "invoice_no": None,
                         "issue_date": None, "vendor": None, "subtotal": None,
                         "tax": None, "total": None, "rows": [],
                         "checks": [f"読み取りに失敗: {e}"]})
    write_excel(recs, a.out)
    if a.json:
        Path(a.json).write_text(json.dumps(recs, ensure_ascii=False, indent=1),
                                encoding="utf-8")
    ng = sum(1 for r in recs if r["checks"])
    print(f"{len(recs)}枚を読み込み、{ng}枚が要確認。→ {a.out}")

if __name__ == "__main__":
    main()
